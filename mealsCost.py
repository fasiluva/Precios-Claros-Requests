# Programa hecho por Valentin Sosa, vea mas de mis trabajos en:
# GitHub: https://github.com/fasiluva 
# Linkedin: https://www.linkedin.com/in/valentin-sosa-aa55a9294/.

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill 
# from pprint import pprint

"""
. El precio por unidad del producto se calcula automaticamente en el excel. No hay que programarlo.
. Hay recetas que usan otras recetas para elaborar productos (por ejemplo, el pan). Estos no son encontrados,
  ya que no estan en el excel de productos. 

"""

def extraerPrecios(sheet : Workbook) -> dict[str, list[float, float]]:
    """
    * Extrae todos los productos de la lista del excel. Guarda el nombre del mismo y
      el precio estandarizado dividido la cantidad de unidades en el.

    Por ejemplo, en 1 kg de anchoas hay aproximadamente 65 anchoas, y el precio del kilo de anchoas 
    sale 20.000, entonces el precio de cada anchoa ronda los 300 pesos. En este caso, el precio de cada 
    anchoa es lo que nos interesa.
    """

    productos = {}

    for row in sheet.iter_rows(min_row = 7, min_col = 2, max_col = 9):
        if row[0].value != None:
            # Si la celda del nombre del producto no es None, entonces no termino la lista.
            try: 
                # print(row[0].value)
                productos[row[0].value.casefold()] = [float(row[6].value), float(row[7].value)]

            except (TypeError, ValueError):
                # Si la casilla de "Precio (por kg/lt/un)" tiene un '?', no lo agrega. 
                # Si la casilla "Agrupados en" no es un numero, no lo agrega.
                continue
        
        else:
            # Si la celda es None, termino la lista.
            break

    return productos


def main(): 

    # Ruta del archivo
    archivo_xlsx = 'Costodecarta(2).xlsx'

    try: 
        workbook = load_workbook(filename=archivo_xlsx)
        sheet = workbook.get_sheet_by_name("Maestro de provedores")

    except:
        print("ERROR: Archivo de materias primas no encontrado. Contactar con Valentin!.\n\nEnter para salir.")
        s = input()
        return

    productos = extraerPrecios(sheet)
    sheet = workbook["CARTA"]
    columnaNombres = sheet['D']

    iterRow = 6     #! Desfase de la hoja del Excel (MODIFICAR SEGUN EXCEL, SI ES NECESARIO)

    while True: 
        try: 
            if columnaNombres[iterRow].fill.patternType == 'solid':
                # Si la celda tiene un subrayado, entonces esta celda es el nombre del platillo.

                iterRow += 1

                while columnaNombres[iterRow].value != None:
                    # Mientras que la celda en la que estoy no sea None, entonces esta es el nombre de un producto.

                    try: 
                        # Intenta buscar el producto en la lista del excel.
                        productoActual = productos[columnaNombres[iterRow].value.casefold()]

                    except KeyError:
                        # Si no encontro el producto, lo notifica, cambia el color de la celda y continua a la siguiente iteracion.
                        print("\nNO ENCONTRADO: ", columnaNombres[iterRow].value.casefold())
                        sheet[f'E{iterRow + 1}'].fill = PatternFill(fill_type='solid', fgColor='E6B8B7')
                        iterRow += 1
                        continue 
                        
                    else: 
                        # Si encontro el producto en la lista del excel:
                        
                        # print(productoActual) 

                        try: 
                            # Intenta escribir en la casilla el precio del producto estandarizado.
                            
                            if sheet[f'C{iterRow + 1}'].value != 'un' and sheet[f'C{iterRow + 1}'].value != 'u':
                                sheet[f'E{iterRow + 1}'] = productoActual[0]

                            else: 
                                # Si la unidad esta medida en unidades, el precio se evalua devuelve en unidades, para que el calculo
                                # sea correcto.
                                sheet[f'E{iterRow + 1}'] = productoActual[0] / productoActual[1]
                            
                            sheet[f'E{iterRow + 1}'].fill = PatternFill(fill_type= 'solid', fgColor= 'B8CCE4') # Color azul

                        except: 
                            # Si no pudo escribir el precio estandarizado, marca la casilla como incompleta y continua la siguiente iteracion.
                            sheet[f'E{iterRow + 1}'].fill = PatternFill(fill_type='solid', fgColor='E6B8B7') # Color rojo
                            iterRow += 1
                            continue
                    
                    iterRow += 1

        except IndexError:
            # Si hubo un IndexError, se llego al final del excel y no hay mas recetas. Sale del while.
            break
        
        else: 
            iterRow += 1

    workbook.save(filename=archivo_xlsx)


main()
print("\nListo!. Enter para salir.")
input()