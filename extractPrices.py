from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import requests
from datetime import datetime
# from pprint import pprint

"""
CADA PRODUCTO TIENE LA SIGUIENTE FORMA (en el ejemplo, se busco "galletitas oreo" en el request): 
{
    "marca": "OREO", 
    "id": "7622201521967", 
    "precioMax": 450.53, 
    "precioMin": 450.53, 
    "nombre": "Galletitas Oreo 36 Un 118 Gr", 
    "presentacion": "118.0 gr", 
    "cantSucursalesDisponible": 1,
    ... 
}
"""


def precioEstandarizado(cantidad : float, unidad : str, precio : float) -> float | str:
    """
    * Trata de pasar las unidades de masa a kg, las de unidades a una unidad y las volumetricas a lt. Si no encuentra la unidad, devuelve '?'.
      Simplemente esta haciendo una regla de 3 simple. Por ejemplo:
      
    cantidad = 150
    unidad = 'gr'
    precio = 450

    150 gr  --> 450 pesos 
    1000 gr --> x

    x = 1000 * 450 / 150 = 3000 pesos.  
    """

    unidadesMayores = {'lt', 'kg', 'u', 'un', 'kgm', 'uni', 'unidad', 'lts', 'l'}
    unidadesMenores = {'cc', 'gr', 'ml', 'mlt', 'grs', 'g', 'mm', 'gramos'}

    if unidad.casefold() in unidadesMayores:
        return precio / cantidad
    
    if unidad.casefold() in unidadesMenores:
        return precio / (cantidad / 1000)

    else: 
        return '?'


def promedio(min : int, max : int) -> float:
    """
    * Calcula el promedio entre los 2 valores.
    """

    return (min + max) / 2


def extract_products(sheet : Workbook) -> list[str]:
    """
    * Recibe la hoja de calculo y extrae los productos.
    * Si la marca del producto es NA, quiere decir que no sera buscado.
    * Si la marca del producto no es NA, entonces sera buscada.

    ! NO PUEDE HABER ESPACIOS EN BLANCO ENTRE PRODUCTOS !
    """

    nombres = sheet['B']    # Lista de nombres de productos
    marca = sheet['C']      # Lista de marcas
    nombresFinal = []

    # Variable que indica desde que fila se empieza a contar los productos
    desfaseArch = 6
    iter = 0

    while True:
        
        try: 

            if marca[desfaseArch + iter].value == 'NA' or marca[desfaseArch + iter].value == "ANCLADO":
                nombresFinal.append('-')

            else: 
                nombresFinal.append(nombres[desfaseArch + iter].value)
            
            iter += 1
                
        except IndexError: # Ya no hay mas productos que buscar
            break

    return nombresFinal


def main():

    # Direccion relativa al archivo
    archivo_xlsx = 'Costodecarta(2).xlsx'

    try: 
        workbook = load_workbook(filename=archivo_xlsx)
        sheet = workbook["Maestro de provedores"]
    except:
        print("ERROR: Archivo de materias primas no encontrado. Contactar con Valentin!.\n\nEnter para salir.")
        s = input()
        return
    else: 
        productos = extract_products(sheet)

    # Headers que se usaran en la request. No tocar!
    headers = {
        'Method': 'GET',
        'Accept': 'application/json, text/plain, */*',
        'Referer': 'https://www.preciosclaros.gob.ar/',
        'X-Api-Key': "zIgFou7Gta7g87VFGL9dZ4BEEs19gNYS1SOQZt96",
        'Accept-Language': 'es-ES,es;q=0.9',
        "Sec-Ch-Ua": '"Chromium";v="116", "Not)A;Brand";v="24", "Opera GX";v="102"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 OPR/102.0.0.0"
    }

    i = 0
    desfaseArch = 7 

    for producto in productos:

        # Si la marca del producto no es 'NA' := '-'
        if producto != '-':

            # URL de la peticion
            url = f"https://d3e6htiiul5ek9.cloudfront.net/prod/productos?string={producto}&array_sucursales=15-1-5274,15-1-5169,19-1-00616,13-1-126,15-1-280,10-2-256,2005-1-49,10-1-156,24-1-337,15-1-6017,15-1-211,9-1-631,2005-1-41,15-1-5180,19-1-02064,23-1-6294,15-1-282,15-1-5482,15-1-5586,15-1-5219,15-1-243,2-1-100,10-1-229,11-4-1046,24-1-317,15-1-234,2005-1-24,15-1-5575,10-2-255,9-1-669&offset=0&limit=50&sort=-cant_sucursales_disponible"

            response = requests.get(url, headers= headers)

            if response.status_code != 200: 
                print("ERROR: Request denegada. Contactar con Valentin!. Estado: ", response.status_code, "/n/nEnter para salir")
                s = input()
                return

            # Pasa el response a un diccionario.
            response = response.json()

            print(f"\n\nProducto: {producto}\n")

            try: 
                # Si existe la categoria 'agrupables', entonces el precio se promedia.
                precioPromedio = promedio(response['agrupables'][0]['precioMin'], response['agrupables'][0]['precioMax'])
                sheet[f'F{desfaseArch + i}'] = precioPromedio
                sheet[f'H{desfaseArch + i}'] = precioPromedio

            except:
                # Si el producto no tiene una agrupacion, entonces el precio se tomara del primer producto que encuentre.

                if response['productos'] != []:
                    # Si la lista de productos encontrados no es vacio:

                    # Separa la en cantidad y unidad usada. Ejemplo: 110.5 gr => (110.5, 'gr')
                    cantidad = response['productos'][0]['presentacion'].split()

                    precio = float(response['productos'][0]['precioMax'])

                    sheet[f'C{desfaseArch + i}'] = response['productos'][0]['marca']                                     # Escribe la marca
                    sheet[f'C{desfaseArch + i}'].fill = PatternFill(fill_type= None)                                     # Quita el estilo previo
                    sheet[f'E{desfaseArch + i}'] = cantidad[1]                                                           # Escribe la unidad usada.
                    sheet[f'F{desfaseArch + i}'] = precio                                                                # Escribe el precio
                    sheet[f'G{desfaseArch + i}'] = f'{datetime.now().day}/{datetime.now().month}/{datetime.now().year}'  # Escribe la fecha de actualizacion
                    
                    try: 
                        # Intenta escribir la cantidad del producto y luego calcular el precio por kg, lt o unidad, segun corresponda.
                        sheet[f'D{desfaseArch + i}'] = float(cantidad[0])
                        sheet[f'H{desfaseArch + i}'] = precioEstandarizado(float(cantidad[0]), cantidad[1], precio)

                    except: 
                        # No pudo generalizar el precio y continua a la siguiente iteracion.
                        i += 1 
                        continue

                    print(response['productos'][0]['marca'])
                    print(response['productos'][0]['presentacion'])
                    print(response['productos'][0]['precioMax'])
                
                else: 
                    # La busqueda fallo y no se encontro ningun producto.
                    sheet[f'C{desfaseArch + i}'] = 'NA'
                    sheet[f'C{desfaseArch + i}'].fill = PatternFill(fill_type= 'solid', fgColor='FFFF00')  

            else: 
                # Else del try. Si entra aca, se encontro la categoria 'agrupables' y promedia el precio.
                sheet[f'C{desfaseArch + i}'] = 'PROMEDIO'
                sheet[f'C{desfaseArch + i}'].fill = PatternFill(fill_type= 'solid', fgColor='FCD5B4')
                sheet[f'G{desfaseArch + i}'] = f'{datetime.now().day}/{datetime.now().month}/{datetime.now().year}'
                
        i += 1 

    workbook.save(filename=archivo_xlsx)    # Guarda el archivo. Si se sale antes, no guardara nada.
    workbook.close()


main()

"""
headers = {
        'authority': 'd3e6htiiul5ek9.cloudfront.net',
        'method': '',
        'path': '/prod/productos?&id_categoria=01&array_sucursales=15-1-5274,15-1-5169,19-1-00616,13-1-126,15-1-280,10-2-256,2005-1-49,10-1-156,24-1-337,15-1-6017,15-1-211,9-1-631,2005-1-41,15-1-5180,19-1-02064,23-1-6294,15-1-282,15-1-5482,15-1-5586,15-1-5219,15-1-243,2-1-100,10-1-229,24-1-317,15-1-234,2005-1-24,15-1-5575,10-2-255,9-1-669,15-1-5582&offset=0&limit=50&sort=-cant_sucursales_disponible',
        'scheme': 'https',
        'Acces-Control-Request-Method': 'GET',
        'Accept-Language': 'es-ES,es;q=0.9',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 OPR/102.0.0.0',
    }
"""
